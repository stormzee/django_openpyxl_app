from django.shortcuts import render, redirect, Http404
# from django.contrib.sessions import sessions
from .forms import UploadFileForm, EmployeeCreationForm
# from django.contrib.auth.forms import UserCreationForm, User
from .models import UploadFile, Employee, Supervisor
import openpyxl

def index(request):
    form = UploadFileForm
    employee_excel_data = []
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_data']
            form.save()
            wkbook = openpyxl.load_workbook(excel_file)
            if wkbook.active:
                wksheet = wkbook.active    
    return render(request, 'index.html', context={'form':form})

def Uploads(request):
    employee_excel_data = []
    excel_files = UploadFile.objects.all()
    for file in excel_files:
        # print(file)
        wrkbook = openpyxl.load_workbook(file.excel_data)
        # let's grab the active worksheet
        wrksheet = wrkbook.active
        for row in wrksheet.iter_rows():
                data = []
                for cell in row:
                    data.append(str(cell.value))
                employee_excel_data.append(data)
    return render(request, 'uploads.htm',context={'employee_excel_data':employee_excel_data})
    
    
def login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        if username == 'admin' and password == 'keypass':
            return redirect('/')
    return render(request, 'login.htm')

def createEmployee(request):
    form = EmployeeCreationForm
    if request.method == 'POST':
        form = EmployeeCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')
    context = {'form':form}
    return render(request, 'Employee.htm', context)

def all_employees(request):
    records = Employee.objects.all()
    context = {'records':records}
    return render(request, 'records.htm', context)
